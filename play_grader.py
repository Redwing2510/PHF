"""
play_grader.py

PFF-style play-level grading derived from AllThreeZones Tracking sheets.

Every tracked event is assigned a grade based on outcome rates derived from
the full dataset (284 games):

  Zone entries  — graded by type (C/D/F/X), lane, pass-play flag, and
                  whether a scoring chance resulted
  Entry defense — graded by what the defender allowed (entry type + chance)
  Zone exits    — graded by exit type (CEX/PEX/CLE/MEX/FEX)
  Shots/chances — graded by SC?, SOG?, G? and primary/secondary assist

Raw play grades are accumulated per player per game as per-60 rates,
z-scored by position group (F vs D) across all tracked games, and
converted to 0-100.  Output interface matches MicrostatGrade so season.py
can use either system transparently.
"""

import re
import pickle
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

from manual_loader import _norm_name as _nn

LOGS_DIR    = Path(__file__).parent / 'Manual Game Logs'
_CACHE_FILE = Path(__file__).parent / '.pg_grades_cache.pkl'

# ---------------------------------------------------------------------------
# Grade constants (outcome rates derived from 284-game dataset)
# ---------------------------------------------------------------------------

_ENTRY_BASE: Dict[Tuple[str, str], float] = {
    ('C', 'C'):  1.0,   # carry center  — 38.9% SC rate
    ('C', 'L'):  0.6,   # carry left    — 22.5%
    ('C', 'R'):  0.6,   # carry right   — 23.8%
    ('D', 'C'):  0.0,   # dump center   — 9.4%  (baseline)
    ('D', 'L'):  0.0,   # dump left     — 8.6%
    ('D', 'R'):  0.0,   # dump right    — 8.4%
    ('F', 'C'): -0.8,   # failed entry
    ('F', 'L'): -0.8,
    ('F', 'R'): -0.8,
    ('X', 'C'):  0.3,   # exchange/contested
    ('X', 'L'):  0.3,
    ('X', 'R'):  0.3,
}
_PASS_BONUS   =  0.2
_CHANCE_BONUS =  0.4
_NO_CHANCE    = -0.1

_DEF_GRADE: Dict[Tuple[str, bool], float] = {
    ('F', False):  0.8,
    ('F', True):   0.8,
    ('X', False):  0.3,
    ('X', True):   0.1,
    ('C', False):  0.1,
    ('C', True):  -0.4,
    ('D', False):  0.0,
    ('D', True):  -0.2,
}

_EXIT_GRADE: Dict[str, float] = {
    'CEX':  0.5,
    'PEX':  0.4,
    'CLE':  0.1,
    'MEX': -0.2,
    'FEX': -0.5,
}

_SHOT_GOAL =  1.5
_SHOT_SC   =  0.8
_SHOT_SOG  =  0.3
_A1_GOAL   =  1.2
_A1_SC     =  0.7
_A2_GOAL   =  0.3
_A2_SC     =  0.175

_CAT_W = {
    'F': {'entries': 0.30, 'defense': 0.25, 'exits': 0.20, 'shots': 0.25},
    'D': {'entries': 0.20, 'defense': 0.35, 'exits': 0.30, 'shots': 0.15},
}

_TEAM_NORM = {'T.B': 'TBL', 'L.A': 'LAK', 'N.J': 'NJD', 'S.J': 'SJS'}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _nt(t: str) -> str:
    return _TEAM_NORM.get(t, t)


def _ss(v) -> str:
    return str(v).strip().upper() if v is not None else ''


def _clamp(z: float) -> float:
    return max(-2.0, min(2.0, z))


def _to100(g: float) -> float:
    return round((g + 2.0) / 4.0 * 100.0, 1)


def _p60(s: float, n: int, toi_min: float) -> float:
    if toi_min <= 0 or n == 0:
        return 0.0
    return s / (toi_min / 60.0)


def _mean_std(vals: list) -> Tuple[Optional[float], Optional[float]]:
    n = len(vals)
    if n < 3:
        return None, None
    mean = sum(vals) / n
    var  = sum((x - mean) ** 2 for x in vals) / (n - 1)
    return mean, (math.sqrt(var) if var > 1e-18 else None)


def _zgrade(val: float, mean: Optional[float], std: Optional[float]) -> float:
    if mean is None or std is None:
        return 0.0
    return _clamp((val - mean) / std)


def _leverage(period, state) -> float:
    """Leverage multiplier: late, close-game plays count more."""
    try:
        p = int(period) if period is not None else 1
        s = abs(int(state)) if state is not None else 0
    except (TypeError, ValueError):
        return 1.0
    if p >= 4:
        return 1.6
    if p == 3:
        if s == 0: return 1.4
        if s == 1: return 1.2
        if s >= 3: return 0.7
    return 1.0


def _season_year_from_path(path: Path) -> int:
    """Infer season start year from folder structure (e.g. '2024-25' → 2024)."""
    for part in path.parts:
        m = re.match(r'^(\d{4})-\d{2}$', part)
        if m:
            return int(m.group(1))
    return 2025  # default for Playoffs/


def _parse_code(code) -> Tuple[Optional[int], Optional[str]]:
    """Parse '38CHI' or '91T.B' -> (jersey_int, team_abbrev)."""
    if not code:
        return None, None
    s = str(code).strip()
    if s.upper() in ('N', '', 'NONE'):
        return None, None
    m = re.match(r'^(\d+)([A-Za-z.]+)$', s)
    if not m:
        return None, None
    return int(m.group(1)), _nt(m.group(2).upper())


# ---------------------------------------------------------------------------
# Per-player accumulator
# ---------------------------------------------------------------------------

class _Acc:
    __slots__ = ('pos', 'toi_min',
                 'ent_s', 'ent_n',
                 'def_s', 'def_n',
                 'ext_s', 'ext_n',
                 'sht_s', 'sht_n',
                 'p_entries', 'p_carries', 'p_pass_c',
                 'p_exits',   'p_ctrl_ext')

    def __init__(self, pos: str, toi: float):
        self.pos     = pos
        self.toi_min = float(toi)
        self.ent_s = 0.0;  self.ent_n = 0
        self.def_s = 0.0;  self.def_n = 0
        self.ext_s = 0.0;  self.ext_n = 0
        self.sht_s = 0.0;  self.sht_n = 0
        self.p_entries  = 0;  self.p_carries  = 0;  self.p_pass_c = 0
        self.p_exits    = 0;  self.p_ctrl_ext = 0


@dataclass
class PlayGrade:
    """Play-level grade for one player in one game, 0-100 scale."""
    offense_100:      float = 50.0
    entries_100:      float = 50.0
    exits_100:        float = 50.0
    defense_100:      float = 50.0
    poss_100:         float = 50.0
    forechecking_100: float = 50.0   # not tracked at play level; neutral
    overall_100:      float = 50.0
    position:         str   = 'F'
    toi_min:          float = 0.0
    plays_graded:     int   = 0


# ---------------------------------------------------------------------------
# Single-file loader
# ---------------------------------------------------------------------------

def _load_one(path: Path) -> Optional[Tuple[int, Dict[str, _Acc]]]:
    m = re.match(r'^(\d+)', path.name)
    if not m:
        return None
    file_id    = int(m.group(1))
    season_year = _season_year_from_path(path)
    game_id    = int(f"{season_year}0{file_id:05d}")

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return None

    pmap: Dict[Tuple[int, str], Tuple[str, str, float]] = {}
    if 'Player List' in wb.sheetnames:
        for row in wb['Player List'].iter_rows(min_row=2, values_only=True):
            name = row[1];  team = row[2];  pos = row[3];  toi = row[4]
            try:
                jersey = int(row[0])
            except (TypeError, ValueError):
                continue
            if not name or not isinstance(name, str):
                continue
            try:
                toi_f = float(toi) if toi is not None else 0.0
            except (TypeError, ValueError):
                toi_f = 0.0
            if toi_f <= 0:
                continue
            ts = _nt(str(team).strip().upper()) if team else ''
            ps = 'D' if str(pos).strip().upper() == 'D' else 'F'
            pmap[(jersey, ts)] = (_nn(name), ps, toi_f)

    if not pmap:
        return None

    acc: Dict[str, _Acc] = {}
    for (_, _), (nm, ps, toi) in pmap.items():
        if nm not in acc:
            acc[nm] = _Acc(ps, toi)

    def _lookup(jersey: int, team: str) -> Optional[_Acc]:
        entry = pmap.get((jersey, team))
        return acc.get(entry[0]) if entry else None

    def _ga_code(code) -> Optional[_Acc]:
        j, t = _parse_code(code)
        return _lookup(j, t) if j is not None else None

    def _ga_jersey(jersey_raw, team_raw) -> Optional[_Acc]:
        if jersey_raw is None:
            return None
        try:
            j = int(float(jersey_raw))
        except (TypeError, ValueError):
            return None
        return _lookup(j, _nt(str(team_raw).strip().upper()) if team_raw else '')

    if 'Tracking' not in wb.sheetnames:
        return (game_id, acc)

    for row in wb['Tracking'].iter_rows(min_row=2, values_only=True):
        entry_type = row[24]
        shooter    = row[4]
        exit_by    = row[34]
        lev        = _leverage(row[0], row[20])

        # ── Shot event ──────────────────────────────────────────────────────
        if shooter is not None and entry_type is None and exit_by is None:
            team = row[3]
            if not team:
                continue
            sc   = _ss(row[12]) == 'Y'
            sog  = _ss(row[13]) == 'Y'
            goal = _ss(row[19]) == 'Y'

            if goal and sc:
                sg = _SHOT_GOAL
            elif sc:
                sg = _SHOT_SC
            elif sog:
                sg = _SHOT_SOG
            else:
                sg = 0.0

            s_acc = _ga_jersey(shooter, team)
            if s_acc and sg > 0:
                s_acc.sht_s += sg * lev;  s_acc.sht_n += 1

            if goal or sc:
                a1g = _A1_GOAL if goal else _A1_SC
                a2g = _A2_GOAL if goal else _A2_SC
                a1  = _ga_jersey(row[6], team)
                if a1:
                    a1.sht_s += a1g * lev;  a1.sht_n += 1
                a2  = _ga_jersey(row[7], team)
                if a2:
                    a2.sht_s += a2g * lev;  a2.sht_n += 1

        # ── Entry event ─────────────────────────────────────────────────────
        elif entry_type is not None:
            et = _ss(entry_type)
            if et not in ('C', 'D', 'F', 'X'):
                continue

            lane   = _ss(row[28])
            if lane not in ('L', 'R', 'C'):
                lane = 'R'

            pass_v = _ss(row[27]) == 'Y'
            chance = _ss(row[30]) == 'Y'

            base = _ENTRY_BASE.get((et, lane), 0.0)
            if pass_v:
                base += _PASS_BONUS
            base += _CHANCE_BONUS if chance else _NO_CHANCE

            att = _ga_code(row[25])
            if att:
                att.ent_s += base * lev;  att.ent_n += 1
                att.p_entries += 1
                if et == 'C':
                    att.p_carries += 1
                    if pass_v:
                        att.p_pass_c += 1

            dg  = _DEF_GRADE.get((et, chance), 0.0)
            dfn = _ga_code(row[26])
            if dfn:
                dfn.def_s += dg * lev;  dfn.def_n += 1

        # ── Exit event ───────────────────────────────────────────────────────
        elif exit_by is not None:
            exit_result = _ss(row[35])
            eg = _EXIT_GRADE.get(exit_result)
            if eg is None:
                continue
            ext = _ga_code(exit_by)
            if ext:
                ext.ext_s += eg * lev;  ext.ext_n += 1
                ext.p_exits += 1
                if exit_result in ('CEX', 'PEX'):
                    ext.p_ctrl_ext += 1

    return (game_id, acc)


# ---------------------------------------------------------------------------
# Load all games
# ---------------------------------------------------------------------------

def _load_all_raw() -> List[Tuple[int, Dict[str, _Acc]]]:
    if not _OPENPYXL_AVAILABLE or not LOGS_DIR.exists():
        return []
    files  = sorted(LOGS_DIR.rglob('*.xlsx'))
    total  = len(files)
    result = []
    for i, f in enumerate(files, 1):
        r = _load_one(f)
        if r:
            result.append(r)
        if i % 25 == 0 or i == total:
            print(f'  Play-grading xlsx files... {i}/{total}', flush=True)
    return result


# ---------------------------------------------------------------------------
# Normalization
# ---------------------------------------------------------------------------

def _compute(raw_list: List[Tuple[int, Dict[str, _Acc]]]) -> Dict[Tuple[int, str], 'PlayGrade']:
    flat   = [(fid, nm, a) for fid, acc_map in raw_list for nm, a in acc_map.items()]
    result: Dict[Tuple[int, str], PlayGrade] = {}

    for pos in ('F', 'D'):
        pool = [(fid, nm, a) for fid, nm, a in flat if a.pos == pos]
        if not pool:
            continue

        # Compute per-60 values once per player; build pools from players with events
        p60: Dict[str, Dict[Tuple, float]] = {cat: {} for cat in ('ent', 'def', 'ext', 'sht', 'off', 'cr', 'er', 'pr')}

        for fid, nm, a in pool:
            key = (fid, nm)
            if a.ent_n > 0:
                p60['ent'][key] = _p60(a.ent_s, a.ent_n, a.toi_min)
            if a.def_n >= 2:  # ≥2 events for a meaningful defense grade (1 event is too noisy)
                p60['def'][key] = _p60(a.def_s, a.def_n, a.toi_min)
            if a.ext_n > 0:
                p60['ext'][key] = _p60(a.ext_s, a.ext_n, a.toi_min)
            if a.sht_n > 0:
                p60['sht'][key] = _p60(a.sht_s, a.sht_n, a.toi_min)
            if a.ent_n + a.sht_n > 0:
                p60['off'][key] = _p60(a.ent_s + a.sht_s, a.ent_n + a.sht_n, a.toi_min)
            # Possession rates (raw 0–1, z-scored like per-60 values)
            if a.p_entries >= 2:
                p60['cr'][key] = a.p_carries / a.p_entries
            if a.p_exits >= 2:
                p60['er'][key] = a.p_ctrl_ext / a.p_exits
            if a.p_carries >= 2:
                p60['pr'][key] = a.p_pass_c / a.p_carries

        # Precompute (mean, std) once per category pool
        stats = {cat: _mean_std(list(vals.values())) for cat, vals in p60.items()}

        w = _CAT_W[pos]
        poss_w = {'F': {'cr': 0.40, 'er': 0.30, 'pr': 0.30},
                  'D': {'cr': 0.25, 'er': 0.50, 'pr': 0.25}}[pos]

        for fid, nm, a in pool:
            key = (fid, nm)

            def _g(cat):
                v = p60[cat].get(key)
                return _zgrade(v, *stats[cat]) if v is not None else 0.0

            ent_g = _g('ent')
            def_g = _g('def')
            ext_g = _g('ext')
            sht_g = _g('sht')
            off_g = _g('off')

            poss_g = _clamp(
                poss_w['cr'] * _g('cr') +
                poss_w['er'] * _g('er') +
                poss_w['pr'] * _g('pr')
            )

            ov = _clamp(
                w['entries'] * ent_g +
                w['defense'] * def_g +
                w['exits']   * ext_g +
                w['shots']   * sht_g
            )

            result[key] = PlayGrade(
                offense_100  = _to100(off_g),
                entries_100  = _to100(ent_g),
                exits_100    = _to100(ext_g),
                defense_100  = _to100(def_g),
                poss_100     = _to100(poss_g),
                overall_100  = _to100(ov),
                position     = pos,
                toi_min      = a.toi_min,
                plays_graded = a.ent_n + a.def_n + a.ext_n + a.sht_n,
            )

    return result


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

_CACHE: Optional[Dict] = None


def _xlsx_fingerprint() -> float:
    mtimes = [f.stat().st_mtime for f in LOGS_DIR.rglob('*.xlsx')]
    return max(mtimes) if mtimes else 0.0


def _build_season_def(raw_list: List[Tuple[int, Dict[str, '_Acc']]]) -> Dict:
    """Aggregate def_s, def_n, toi_min per (season_year, player_name) across all games."""
    agg: Dict = {}
    for game_id, acc_map in raw_list:
        sy = int(str(game_id)[:4])
        for nm, a in acc_map.items():
            k = (sy, nm)
            if k not in agg:
                agg[k] = [0.0, 0, 0.0, a.pos]
            agg[k][0] += a.def_s
            agg[k][1] += a.def_n
            agg[k][2] += a.toi_min
    return {k: tuple(v) for k, v in agg.items()}


def load_play_grades() -> Dict:
    global _CACHE
    if _CACHE is not None:
        return _CACHE

    if _CACHE_FILE.exists() and _xlsx_fingerprint() <= _CACHE_FILE.stat().st_mtime:
        print('  Loading play grades from cache...', flush=True)
        with open(_CACHE_FILE, 'rb') as f:
            loaded = pickle.load(f)
        if '_season_def' in loaded:
            _CACHE = loaded
            grade_count = sum(1 for k in _CACHE if k != '_season_def')
            print(f'  Cache loaded ({grade_count} player-game records).', flush=True)
            return _CACHE
        print('  Play grade cache missing season aggregates, rebuilding...', flush=True)

    raw    = _load_all_raw()
    _CACHE = _compute(raw)
    _CACHE['_season_def'] = _build_season_def(raw)
    with open(_CACHE_FILE, 'wb') as f:
        pickle.dump(_CACHE, f)
    grade_count = sum(1 for k in _CACHE if k != '_season_def')
    print(f'  Play grade cache saved ({grade_count} player-game records).', flush=True)
    return _CACHE


def get_play_grade(game_id: int, name: str) -> Optional[PlayGrade]:
    """
    Look up a play-level grade for a player in a game.

    game_id: full NHL game ID (e.g. 2025030131)
    name:    player name (any casing, non-breaking spaces OK)
    """
    grades = load_play_grades()
    return grades.get((game_id, _nn(name)))


def get_season_def_aggregates(season_year: int) -> Dict[str, Tuple]:
    """Return {norm_name: (def_s, def_n, toi_min, pos)} for all tracked players in the given season year."""
    cache = load_play_grades()
    season_def = cache.get('_season_def', {})
    return {nm: v for (sy, nm), v in season_def.items() if sy == season_year}


def invalidate_cache() -> None:
    global _CACHE
    _CACHE = None
    if _CACHE_FILE.exists():
        _CACHE_FILE.unlink()

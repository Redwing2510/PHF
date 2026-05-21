"""
manual_loader.py

Reads AllThreeZones-style xlsx game logs from 'Manual Game Logs/' and produces
PFF-style -2 to +2 microstat grades per player per game.

Grades are computed by:
  1. Per-60 normalizing each raw count stat
  2. Z-scoring each metric relative to the same position group (F or D)
     across ALL tracked games in the dataset
  3. Weighting z-scores into 5 category grades: Offense, Entries, Exits,
     Entry Defense, Forechecking
  4. Weighting categories (F/D have different weights) into an overall grade
  5. Clamping all grades to [-2, +2] and converting to 0-100 for display
Should
Usage:
    from manual_loader import get_microstat_grade
    ms = get_microstat_grade(game_id=2025030131, name="Sebastian Aho")
    if ms:
        print(ms.overall, ms.offense, ms.entries)  # -2 to +2
        print(ms.overall_100)                        # 0-100

Call invalidate_cache() after dropping new xlsx files into Manual Game Logs/.
"""

import os
import re
import pickle
import math
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

LOGS_DIR   = Path(__file__).parent / "Manual Game Logs"
_CACHE_FILE = Path(__file__).parent / ".ms_grades_cache.pkl"

# ---------------------------------------------------------------------------
# Column indices in the Player List sheet (0-indexed)
# ---------------------------------------------------------------------------
C_JERSEY    = 0
C_NAME      = 1
C_TEAM      = 2
C_POS       = 3
C_TOI       = 4   # 5v5 TOI in minutes
C_YEAR      = 5
C_GAME      = 6
C_SHOTS     = 7   # Shot attempts
C_SOG       = 8   # Shots On Goal
C_CHANCES   = 9   # Scoring Chances
# C_PASSES  = 10
C_PA1       = 11  # Primary Shot Assists
C_PA2       = 12  # Secondary Shot Assists
C_PA3       = 13  # Tertiary Shot Assists
C_CA        = 14  # Chance Assists
C_HOME_PLATE     = 15   # Home Plate assists
C_LOW_HIGH       = 16   # Low-to-High assists
C_BEHIND_NET     = 17   # Behind Net assists
C_CENTER_LANE    = 18   # Center Lane Assists
C_NZ_ASSIST      = 19   # NZ Assist (pass from NZ)
C_DZ_ASSIST      = 20   # DZ Assist (passer standing in DZ)
C_SHOTS_RUSH     = 21   # Shots off rush
C_ASSISTS_RUSH   = 22   # Assists off Rush
C_SHOTS_FC       = 23   # Shots off Forecheck or Cycle
C_ASSISTS_FC     = 24   # Assists off Forecheck
C_SHOTS_CYCLE    = 25   # Shots off cycle
C_ASSISTS_CYCLE  = 26   # Assists off Cycle
C_SHOTS_HD       = 27   # Shots off HD Passes
C_ENTRIES   = 28  # Zone Entries (total)
C_CARRIES   = 29  # Carries (controlled entries)
C_FAILED_E  = 30  # Failed Entry attempts
C_PASS_E    = 31  # Entries via passing play
C_RECOVERIES     = 32   # OZ Recoveries
C_CWC       = 33  # Carries with Scoring Chances
C_DIC       = 34  # Dump-in Chances
C_FC_PRESS  = 35  # Forecheck Pressures
# C_DZ_TOUCH = 36
C_DZ_RET    = 37  # DZ Retrievals
C_EXITS     = 38  # Zone Exits (total)
C_EWP       = 39  # Exits with Possession
C_CARRY_EXIT = 40  # Carry-outs (controlled exits by skating)
C_PASS_EXIT  = 41  # Pass-exits (controlled exits by passing)
C_CLEARS     = 42  # Clears (dump-outs, any exit under pressure)
C_MISSED_PASS    = 43   # Missed Passes (DZ exit attempt)
C_RET_EXIT  = 44  # Retrievals Leading to Exits
C_BOTCH     = 45  # Botched Retrievals (negative)
C_EXCHANGE  = 46  # Exchanges (contested entries, partial denial credit)
C_FAIL_EXIT = 47  # Failed Exits (negative)
# C_RUSHED    = 48
# C_2ND_TOUCH = 49
C_TARGETS   = 50  # Entry Targets (entries against)
C_CARRIES_AG     = 51   # Carries (entry defense umbrella)
C_DENIALS   = 52  # Carry Denials
C_PASSES_AG = 53  # Passes Allowed (entries against)
C_CCA       = 54  # Carries with Chance Against (negative)
C_DICA      = 55  # Dump-in with Chance Against (negative)
C_OT_SHOTS       = 69   # One-timer shots
C_REBOUNDS       = 70   # Rebound shots
C_DEFLECTIONS    = 71   # Deflection shots
C_OT_ASSIST      = 72   # One-timer Assists
C_REB_ASSIST     = 73   # Rebound Assists
C_DEFLECT_ASSIST = 74   # Deflection Assists
C_NZ_ASSISTS     = 79   # NZ Assists (play originated from NZ)
C_DZ_ASSISTS     = 80   # DZ Assists (play originated from DZ)


@dataclass
class MicrostatRecord:
    """Raw per-player stats extracted from one game log xlsx."""
    game_file_id: int         # short ID from filename, e.g. 30131
    name: str                 # normalized lowercase with spaces
    team: str
    position: str             # 'F' or 'D'
    toi_min: float            # 5v5 TOI in minutes

    # Offense
    shots: int = 0
    sog: int = 0
    chances: int = 0
    primary_assists: int = 0
    secondary_assists: int = 0
    tertiary_assists: int = 0
    chance_assists: int = 0
    shots_off_rush: int = 0
    shots_off_forecheck: int = 0
    shots_off_cycle: int = 0

    # Zone entries
    zone_entries: int = 0
    carries: int = 0
    failed_entries: int = 0
    pass_entries: int = 0
    carries_w_chances: int = 0
    dump_in_chances: int = 0

    # Zone exits
    zone_exits: int = 0
    exits_w_possession: int = 0
    carry_exits: int = 0
    pass_exits: int = 0
    clears: int = 0
    retrievals_leading_to_exits: int = 0
    botched_retrievals: int = 0
    failed_exits: int = 0

    # Entry defense
    targets: int = 0
    denials: int = 0
    pk_denials: int = 0        # kept for backwards compatibility (default 0)
    exchanges: int = 0
    passes_allowed: int = 0
    carries_chance_against: int = 0
    dump_in_chance_against: int = 0

    # Forechecking
    fc_pressures: int = 0
    dz_retrievals: int = 0
    dz_controlled_breakout: int = 0   # kept for backwards compatibility (default 0)

    # Passing — chance assist subtypes
    home_plate_assists: int = 0
    low_high_assists: int = 0
    behind_net_assists: int = 0
    center_lane_assists: int = 0
    assists_off_rush: int = 0
    assists_off_forecheck: int = 0
    assists_off_cycle: int = 0
    onetimer_assists: int = 0
    rebound_assists: int = 0
    deflect_assists: int = 0
    nz_assist: int = 0       # passer was in NZ (col 19)
    dz_assist: int = 0       # passer was in DZ (col 20)
    nz_assists: int = 0      # play originated from NZ (col 79)
    dz_assists: int = 0      # play originated from DZ (col 80)

    # Shooting — shot types
    onetimers: int = 0
    rebounds: int = 0
    deflections: int = 0
    shots_off_hd: int = 0

    # Goals (computed from Tracking sheet)
    goals: int = 0
    primary_goal_assists: int = 0
    secondary_goal_assists: int = 0
    tertiary_goal_assists: int = 0

    # OZ Activity
    recoveries: int = 0

    # DZ Exit
    missed_passes: int = 0

    # Entry Defense
    carries_against: int = 0   # entry defense umbrella


@dataclass
class MicrostatGrade:
    """PFF-style grades for a single player in a single game."""
    # -2 to +2 category grades
    offense: float = 0.0
    entries: float = 0.0
    exits: float = 0.0
    entry_defense: float = 0.0
    forechecking: float = 0.0
    overall: float = 0.0

    # 0-100 equivalents for display alongside API grades
    overall_100: float = 50.0
    offense_100: float = 50.0
    entries_100: float = 50.0
    exits_100: float = 50.0
    defense_100: float = 50.0
    forechecking_100: float = 50.0

    position: str = 'F'
    toi_min: float = 0.0

    # Raw counts for the newly wired columns (shown in game expanded row)
    raw_secondary_assists: int = 0
    raw_pass_entries: int = 0
    raw_shots_off_rush: int = 0
    raw_shots_off_forecheck: int = 0
    raw_dz_breakout: int = 0


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _norm_name(name: str) -> str:
    """Normalize player name: strip non-breaking spaces, lowercase, remove A3Z 'other ' prefix."""
    n = name.replace('\xa0', ' ').strip().lower()
    if n.startswith('other '):
        n = n[6:]
    return n


def _name_parts(norm: str):
    """Return (first_initial, last_name) from a normalized name, or (None, None)."""
    parts = norm.split()
    if len(parts) < 2:
        return None, None
    last = parts[-1]
    first = parts[0].rstrip('.')  # handles "j." → "j"
    return first[0] if first else None, last


_TEAM_ALIASES = {'L.A': 'LAK', 'N.J': 'NJD', 'S.J': 'SJS', 'T.B': 'TBL'}

def _norm_team(team: str) -> str:
    """Normalize team abbreviation to NHL API format."""
    t = team.strip().upper()
    return _TEAM_ALIASES.get(t, t)


def _safe_int(v) -> int:
    try:
        return int(v) if v is not None else 0
    except (TypeError, ValueError):
        return 0


def _safe_float(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def _pos_group(pos: str) -> str:
    if pos in ('D',):
        return 'D'
    return 'F'


def _per60(count: int, toi_min: float) -> float:
    if toi_min <= 0:
        return 0.0
    return count / toi_min * 60.0


def _rate(num: int, denom: int) -> float:
    if denom <= 0:
        return 0.5  # regress to 50% when no opportunity
    return num / denom


# Cache mean/std per pool list (keyed by list id) for the duration of _compute_grades.
# This avoids re-scanning the same pool thousands of times.
_POOL_STATS_CACHE: dict = {}


def _pool_stats(vals: list):
    """Return (mean, std) for vals, memoised by list identity."""
    k = id(vals)
    if k not in _POOL_STATS_CACHE:
        n = len(vals)
        mean = sum(vals) / n
        var = sum((x - mean) ** 2 for x in vals) / (n - 1)
        _POOL_STATS_CACHE[k] = (mean, math.sqrt(var) if var > 1e-18 else None)
    return _POOL_STATS_CACHE[k]


def _zscore(value: float, vals: list) -> float:
    if len(vals) < 3:
        return 0.0
    mean, std = _pool_stats(vals)
    if std is None:
        return 0.0
    return (value - mean) / std


def _clamp(z: float) -> float:
    """Clamp z-score to -2 … +2."""
    return max(-2.0, min(2.0, z))


def _to_100(grade: float) -> float:
    """Convert -2…+2 grade to 0–100."""
    return round((grade + 2.0) / 4.0 * 100.0, 1)


# ---------------------------------------------------------------------------
# xlsx loading
# ---------------------------------------------------------------------------

def _load_records_from_file(path: Path, file_game_id: int) -> list:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if 'Player List' not in wb.sheetnames:
        return []

    # Build goal/goal-assist counts from Tracking sheet
    goal_counts = {}  # (jersey_int, team_str) -> [goals, pa1_goals, pa2_goals, pa3_goals]
    if 'Tracking' in wb.sheetnames:
        ws_track = wb['Tracking']
        # Tracking sheet columns (0-based):
        # C=2: situation, D=3: team, E=4: shooter jersey
        # G=6: A1 jersey, H=7: A2 jersey, I=8: A3 jersey, T=19: goal flag
        for trow in ws_track.iter_rows(min_row=2, values_only=True):
            if not trow or len(trow) < 20:
                continue
            situation = trow[2]
            if situation != '5v5':
                continue
            goal_flag = trow[19]
            if str(goal_flag).lower() != 'y':
                continue
            team = str(trow[3]) if trow[3] else ''
            shooter = trow[4]
            a1 = trow[6]
            a2 = trow[7]
            a3 = trow[8]
            if shooter is not None:
                k = (int(shooter), team)
                if k not in goal_counts:
                    goal_counts[k] = [0, 0, 0, 0]
                goal_counts[k][0] += 1
            if a1 is not None:
                k = (int(a1), team)
                if k not in goal_counts:
                    goal_counts[k] = [0, 0, 0, 0]
                goal_counts[k][1] += 1
            if a2 is not None:
                k = (int(a2), team)
                if k not in goal_counts:
                    goal_counts[k] = [0, 0, 0, 0]
                goal_counts[k][2] += 1
            if a3 is not None:
                k = (int(a3), team)
                if k not in goal_counts:
                    goal_counts[k] = [0, 0, 0, 0]
                goal_counts[k][3] += 1

    ws = wb['Player List']
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[C_NAME]
        pos  = row[C_POS]
        toi  = row[C_TOI]

        if not name or not isinstance(name, str):
            continue
        if not pos or pos == 'G':
            continue
        toi_f = _safe_float(toi)
        if toi_f <= 0:
            continue

        jersey_int = _safe_int(row[C_JERSEY])
        team_str = str(row[C_TEAM]) if row[C_TEAM] else ''
        gc = goal_counts.get((jersey_int, team_str), [0, 0, 0, 0])

        rec = MicrostatRecord(
            game_file_id=file_game_id,
            name=_norm_name(name),
            team=team_str,
            position=_pos_group(str(pos)),
            toi_min=toi_f,
            shots=_safe_int(row[C_SHOTS]),
            sog=_safe_int(row[C_SOG]),
            chances=_safe_int(row[C_CHANCES]),
            primary_assists=_safe_int(row[C_PA1]),
            secondary_assists=_safe_int(row[C_PA2]),
            tertiary_assists=_safe_int(row[C_PA3]),
            chance_assists=_safe_int(row[C_CA]),
            shots_off_rush=_safe_int(row[C_SHOTS_RUSH]),
            shots_off_forecheck=_safe_int(row[C_SHOTS_FC]),
            shots_off_cycle=_safe_int(row[C_SHOTS_CYCLE]),
            zone_entries=_safe_int(row[C_ENTRIES]),
            carries=_safe_int(row[C_CARRIES]),
            failed_entries=_safe_int(row[C_FAILED_E]),
            pass_entries=_safe_int(row[C_PASS_E]),
            carries_w_chances=_safe_int(row[C_CWC]),
            dump_in_chances=_safe_int(row[C_DIC]),
            zone_exits=_safe_int(row[C_EXITS]),
            exits_w_possession=_safe_int(row[C_EWP]),
            carry_exits=_safe_int(row[C_CARRY_EXIT]),
            pass_exits=_safe_int(row[C_PASS_EXIT]),
            clears=_safe_int(row[C_CLEARS]),
            retrievals_leading_to_exits=_safe_int(row[C_RET_EXIT]),
            botched_retrievals=_safe_int(row[C_BOTCH]),
            failed_exits=_safe_int(row[C_FAIL_EXIT]),
            targets=_safe_int(row[C_TARGETS]),
            denials=_safe_int(row[C_DENIALS]),
            exchanges=_safe_int(row[C_EXCHANGE]),
            passes_allowed=_safe_int(row[C_PASSES_AG]),
            carries_chance_against=_safe_int(row[C_CCA]),
            dump_in_chance_against=_safe_int(row[C_DICA]),
            fc_pressures=_safe_int(row[C_FC_PRESS]),
            dz_retrievals=_safe_int(row[C_DZ_RET]),
            home_plate_assists=_safe_int(row[C_HOME_PLATE]),
            low_high_assists=_safe_int(row[C_LOW_HIGH]),
            behind_net_assists=_safe_int(row[C_BEHIND_NET]),
            center_lane_assists=_safe_int(row[C_CENTER_LANE]),
            assists_off_rush=_safe_int(row[C_ASSISTS_RUSH]),
            assists_off_forecheck=_safe_int(row[C_ASSISTS_FC]),
            assists_off_cycle=_safe_int(row[C_ASSISTS_CYCLE]),
            onetimer_assists=_safe_int(row[C_OT_ASSIST]),
            rebound_assists=_safe_int(row[C_REB_ASSIST]),
            deflect_assists=_safe_int(row[C_DEFLECT_ASSIST]),
            nz_assist=_safe_int(row[C_NZ_ASSIST]),
            dz_assist=_safe_int(row[C_DZ_ASSIST]),
            nz_assists=_safe_int(row[C_NZ_ASSISTS]),
            dz_assists=_safe_int(row[C_DZ_ASSISTS]),
            onetimers=_safe_int(row[C_OT_SHOTS]),
            rebounds=_safe_int(row[C_REBOUNDS]),
            deflections=_safe_int(row[C_DEFLECTIONS]),
            shots_off_hd=_safe_int(row[C_SHOTS_HD]),
            goals=gc[0],
            primary_goal_assists=gc[1],
            secondary_goal_assists=gc[2],
            tertiary_goal_assists=gc[3],
            recoveries=_safe_int(row[C_RECOVERIES]),
            missed_passes=_safe_int(row[C_MISSED_PASS]),
            carries_against=_safe_int(row[C_CARRIES_AG]),
        )
        records.append(rec)
    return records


def _load_all_records() -> list:
    if not _OPENPYXL_AVAILABLE or not LOGS_DIR.exists():
        return []
    all_records = []
    files = sorted(LOGS_DIR.rglob('*.xlsx'))
    total = len(files)
    for i, fname in enumerate(files, 1):
        m = re.match(r'^(\d+)', fname.name)
        if not m:
            continue
        file_game_id = int(m.group(1))
        try:
            recs = _load_records_from_file(fname, file_game_id)
            all_records.extend(recs)
        except Exception:
            pass
        if i % 25 == 0 or i == total:
            print(f"  Loading xlsx files... {i}/{total}", flush=True)
    return all_records


# ---------------------------------------------------------------------------
# Grade engine
# ---------------------------------------------------------------------------

# Category weights per position group
# Must sum to 1.0 for each position
_CAT_WEIGHTS = {
    'F': {
        'offense':      0.45,
        'entries':      0.20,
        'defense':      0.35,
    },
    'D': {
        'offense':       0.20,
        'entries':       0.15,
        'exits':         0.30,
        'entry_defense': 0.35,
    },
}


def _compute_grades(records: list) -> dict:
    """
    Grade all records relative to position-group baselines.
    Returns dict keyed by (game_file_id, name_normalized) -> MicrostatGrade.
    """
    _POOL_STATS_CACHE.clear()
    fwd_recs = [r for r in records if r.position == 'F']
    def_recs = [r for r in records if r.position == 'D']

    result = {}

    for pos_recs in [fwd_recs, def_recs]:
        if not pos_recs:
            continue

        pos = pos_recs[0].position

        # Build per-60 and rate vectors for the entire position pool
        shots_p60    = [_per60(r.shots,              r.toi_min) for r in pos_recs]
        chances_p60  = [_per60(r.chances,            r.toi_min) for r in pos_recs]
        pa1_p60      = [_per60(r.primary_assists,    r.toi_min) for r in pos_recs]
        pa2_p60      = [_per60(r.secondary_assists,  r.toi_min) for r in pos_recs]
        pa3_p60      = [_per60(r.tertiary_assists,   r.toi_min) for r in pos_recs]
        ca_p60       = [_per60(r.chance_assists,     r.toi_min) for r in pos_recs]
        rush_rate    = [_rate(r.shots_off_rush,      r.shots)   for r in pos_recs]
        fc_shot_rate = [_rate(r.shots_off_forecheck, r.shots)   for r in pos_recs]

        carry_p60      = [_per60(r.carries,            r.toi_min) for r in pos_recs]
        pass_entry_p60 = [_per60(r.pass_entries,       r.toi_min) for r in pos_recs]
        entry_eff      = [_rate(r.carries, r.carries + r.failed_entries) for r in pos_recs]
        ctrl_rate      = [_rate(r.carries + r.pass_entries, r.zone_entries) for r in pos_recs]
        cwc_p60        = [_per60(r.carries_w_chances,  r.toi_min) for r in pos_recs]

        exit_ctrl      = [_rate(r.exits_w_possession, r.zone_exits) for r in pos_recs]
        carry_exit_p60 = [_per60(r.carry_exits,        r.toi_min) for r in pos_recs]
        pass_exit_p60  = [_per60(r.pass_exits,         r.toi_min) for r in pos_recs]
        clears_p60     = [_per60(r.clears,             r.toi_min) for r in pos_recs]
        ret_exit_p60   = [_per60(r.retrievals_leading_to_exits, r.toi_min) for r in pos_recs]
        botch_p60      = [_per60(r.botched_retrievals, r.toi_min) for r in pos_recs]
        breakout_p60   = [_per60(r.dz_controlled_breakout, r.toi_min) for r in pos_recs]

        denial_p60  = [_per60(r.denials, r.toi_min) for r in pos_recs]
        denial_rate = [_rate(r.denials, r.targets) for r in pos_recs]
        pkdeny_p60  = [_per60(r.pk_denials,  r.toi_min) for r in pos_recs]
        exchange_p60= [_per60(r.exchanges,   r.toi_min) for r in pos_recs]
        cca_p60     = [_per60(r.carries_chance_against, r.toi_min) for r in pos_recs]
        dica_p60    = [_per60(r.dump_in_chance_against, r.toi_min) for r in pos_recs]

        fc_p60      = [_per60(r.fc_pressures,  r.toi_min) for r in pos_recs]
        dzret_p60   = [_per60(r.dz_retrievals, r.toi_min) for r in pos_recs]

        cw = _CAT_WEIGHTS[pos]

        for i, r in enumerate(pos_recs):
            # ── Offense ──────────────────────────────────────────────────
            if pos == 'D':
                z_off = (
                    0.25 * _zscore(chances_p60[i],    chances_p60)  +
                    0.15 * _zscore(shots_p60[i],      shots_p60)    +
                    0.20 * _zscore(pa1_p60[i],        pa1_p60)      +
                    0.06 * _zscore(pa2_p60[i],        pa2_p60)      +
                    0.03 * _zscore(pa3_p60[i],        pa3_p60)      +
                    0.12 * _zscore(ca_p60[i],         ca_p60)       +
                    0.04 * _zscore(rush_rate[i],      rush_rate)    +
                    0.03 * _zscore(fc_shot_rate[i],   fc_shot_rate) +
                    0.12 * _zscore(fc_p60[i],         fc_p60)
                )
            else:
                z_off = (
                    0.25 * _zscore(chances_p60[i],    chances_p60)  +
                    0.18 * _zscore(shots_p60[i],      shots_p60)    +
                    0.19 * _zscore(pa1_p60[i],        pa1_p60)      +
                    0.07 * _zscore(pa2_p60[i],        pa2_p60)      +
                    0.03 * _zscore(pa3_p60[i],        pa3_p60)      +
                    0.12 * _zscore(ca_p60[i],         ca_p60)       +
                    0.06 * _zscore(rush_rate[i],      rush_rate)    +
                    0.04 * _zscore(fc_shot_rate[i],   fc_shot_rate) +
                    0.06 * _zscore(fc_p60[i],         fc_p60)
                )
            off_g = _clamp(z_off)

            # ── Zone Entries ──────────────────────────────────────────────
            z_ent = (
                0.30 * _zscore(ctrl_rate[i],       ctrl_rate)       +
                0.25 * _zscore(carry_p60[i],        carry_p60)       +
                0.20 * _zscore(cwc_p60[i],          cwc_p60)         +
                0.15 * _zscore(pass_entry_p60[i],   pass_entry_p60)  +
                0.10 * _zscore(entry_eff[i],         entry_eff)
            )
            ent_g = _clamp(z_ent)

            # ── Zone Exits ───────────────────────────────────────────────
            z_ext = (
                0.23 * _zscore(carry_exit_p60[i], carry_exit_p60) +
                0.30 * _zscore(exit_ctrl[i],       exit_ctrl)      +
                0.13 * _zscore(pass_exit_p60[i],   pass_exit_p60)  +
                0.10 * _zscore(ret_exit_p60[i],    ret_exit_p60)   +
                0.10 * _zscore(breakout_p60[i],    breakout_p60)   +
                0.04 * _zscore(clears_p60[i],      clears_p60)     -
                0.10 * _zscore(botch_p60[i],       botch_p60)
            )
            ext_g = _clamp(z_ext)

            # ── Entry Defense ─────────────────────────────────────────────
            # Denials (5v5 + PK) + exchanges (partial denial credit) +
            # denial rate - dangerous carries allowed
            all_denials_p60 = denial_p60[i] + pkdeny_p60[i]
            all_denials_pool = [d + pk for d, pk in zip(denial_p60, pkdeny_p60)]
            # Clamp CCA/DICA z-scores before combining — their per-60 rates are
            # extremely sparse (pool mean ~0.1), so one event in short ice time
            # produces z-scores of 8-10 that would dominate the entire grade.
            z_cca_c  = _clamp(_zscore(cca_p60[i],  cca_p60))
            z_dica_c = _clamp(_zscore(dica_p60[i], dica_p60))
            # Shift weight from volume penalties toward rate metrics —
            # OZ-heavy players face fewer entries so each CCA carries
            # disproportionate weight without this adjustment.
            if pos == 'D':
                # D-men: DZ retrievals count as defensive contribution
                z_def = (
                    0.25 * _zscore(denial_rate[i],   denial_rate)      +
                    0.20 * _zscore(all_denials_p60,  all_denials_pool) +
                    0.15 * _zscore(exchange_p60[i],  exchange_p60)     +
                    0.20 * _zscore(dzret_p60[i],     dzret_p60)        -
                    0.15 * z_cca_c                                      -
                    0.05 * z_dica_c
                )
            else:
                z_def = (
                    0.30 * _zscore(denial_rate[i],   denial_rate)      +
                    0.25 * _zscore(all_denials_p60,  all_denials_pool) +
                    0.15 * _zscore(exchange_p60[i],  exchange_p60)     -
                    0.20 * z_cca_c                                      -
                    0.10 * z_dica_c
                )
            def_g = _clamp(z_def)

            # ── Forechecking ──────────────────────────────────────────────
            z_fc = (
                0.50 * _zscore(fc_p60[i],    fc_p60)    +
                0.50 * _zscore(dzret_p60[i], dzret_p60)
            )
            fc_g = _clamp(z_fc)

            # ── Overall ───────────────────────────────────────────────────
            if pos == 'F':
                # Exits (75%) + entry defense (15%) + DZ retrievals (10%)
                dzret_g = _clamp(_zscore(dzret_p60[i], dzret_p60))
                combined_def_g = _clamp(0.75 * ext_g + 0.15 * def_g + 0.10 * dzret_g)
                overall_g = (
                    cw['offense']  * off_g          +
                    cw['entries']  * ent_g          +
                    cw['defense']  * combined_def_g
                )
                display_def_g = combined_def_g
            else:
                overall_g = (
                    cw['offense']       * off_g +
                    cw['entries']       * ent_g +
                    cw['exits']         * ext_g +
                    cw['entry_defense'] * def_g
                )
                display_def_g = def_g
            overall_g = _clamp(overall_g)

            grade = MicrostatGrade(
                offense=round(off_g,          2),
                entries=round(ent_g,          2),
                exits=round(ext_g,            2),
                entry_defense=round(def_g,    2),
                forechecking=round(fc_g,      2),
                overall=round(overall_g,      2),
                overall_100=_to_100(overall_g),
                offense_100=_to_100(off_g),
                entries_100=_to_100(ent_g),
                exits_100=_to_100(ext_g),
                defense_100=_to_100(display_def_g),
                forechecking_100=_to_100(fc_g),
                position=r.position,
                toi_min=r.toi_min,
                raw_secondary_assists=r.secondary_assists,
                raw_pass_entries=r.pass_entries,
                raw_shots_off_rush=r.shots_off_rush,
                raw_shots_off_forecheck=r.shots_off_forecheck,
                raw_dz_breakout=r.dz_controlled_breakout,
            )
            result[(r.game_file_id, r.name)] = grade

    return result


def _xlsx_fingerprint() -> float:
    """Return the most recent mtime across all xlsx files."""
    if not LOGS_DIR.exists():
        return 0.0
    mtimes = [f.stat().st_mtime for f in LOGS_DIR.rglob('*.xlsx')]
    return max(mtimes) if mtimes else 0.0


# ---------------------------------------------------------------------------
# Module-level cache (loaded once per server process)
# Pickle format: {'grades': {(id,name): MicrostatGrade}, 'records': {(id,name): MicrostatRecord}}
# ---------------------------------------------------------------------------
_CACHE:         Optional[dict] = None  # grades dict
_RECORD_CACHE:  Optional[dict] = None  # records dict


def _load_both_caches() -> None:
    """Populate _CACHE and _RECORD_CACHE, updating incrementally for new xlsx files."""
    global _CACHE, _RECORD_CACHE

    existing_grades:  dict = {}
    existing_records: dict = {}
    processed_ids:    set  = set()

    if _CACHE_FILE.exists():
        try:
            with open(_CACHE_FILE, 'rb') as f:
                payload = pickle.load(f)
            if isinstance(payload, dict) and 'grades' in payload:
                existing_grades  = payload['grades']
                existing_records = payload.get('records', {})
                processed_ids    = {k[0] for k in existing_grades}
        except Exception:
            pass

    # Find xlsx files whose game_file_id hasn't been processed yet
    new_files: list = []
    if LOGS_DIR.exists():
        for fname in sorted(LOGS_DIR.rglob('*.xlsx')):
            m = re.match(r'^(\d+)', fname.name)
            if not m:
                continue
            if int(m.group(1)) not in processed_ids:
                new_files.append(fname)

    if not new_files:
        _CACHE        = existing_grades
        _RECORD_CACHE = existing_records
        if _CACHE:
            print(f"  Microstat cache loaded ({len(_CACHE)} player-game records).", flush=True)
        return

    print(f"  Parsing {len(new_files)} new xlsx file(s)...", flush=True)
    new_records = []
    for fname in new_files:
        m = re.match(r'^(\d+)', fname.name)
        file_game_id = int(m.group(1))
        try:
            new_records.extend(_load_records_from_file(fname, file_game_id))
        except Exception:
            pass

    new_grades  = _compute_grades(new_records)
    new_rec_map = {(r.game_file_id, r.name, _norm_team(r.team), r.position): r for r in new_records}

    existing_grades.update(new_grades)
    existing_records.update(new_rec_map)
    _CACHE        = existing_grades
    _RECORD_CACHE = existing_records

    payload = {'grades': _CACHE, 'records': _RECORD_CACHE}
    with open(_CACHE_FILE, 'wb') as f:
        pickle.dump(payload, f)
    print(f"  Microstat cache saved ({len(_CACHE)} player-game records).", flush=True)


def load_microstat_grades() -> dict:
    """
    Returns grades dict keyed by (game_file_id: int, name_normalized: str) -> MicrostatGrade.
    """
    global _CACHE
    if _CACHE is None:
        _load_both_caches()
    return _CACHE


def load_microstat_records() -> dict:
    """
    Returns records dict keyed by (game_file_id: int, name_normalized: str) -> MicrostatRecord.
    """
    global _RECORD_CACHE
    if _RECORD_CACHE is None:
        _load_both_caches()
    return _RECORD_CACHE


def get_microstat_grade(game_id: int, name: str) -> Optional[MicrostatGrade]:
    """
    Look up a microstat grade for a player in a game.

    game_id: full NHL game ID (e.g. 2025030131)
    name:    player name (any casing, non-breaking spaces OK)
    """
    grades = load_microstat_grades()
    norm  = _norm_name(name)
    short = int(str(game_id)[-5:])
    return grades.get((short, norm))


def get_microstat_record(game_id: int, name: str, team: str = '', position: str = '') -> Optional[MicrostatRecord]:
    """
    Look up the raw MicrostatRecord (event counts) for a player in a game.

    game_id:  full NHL game ID (e.g. 2025030131)
    name:     player name (any casing, non-breaking spaces OK)
    team:     team abbreviation — recommended to avoid same-name collisions
    position: player position ('C','L','R','D') — used to disambiguate same-name players on the same team
    """
    records = load_microstat_records()
    norm  = _norm_name(name)
    short = int(str(game_id)[-5:])
    norm_team = _norm_team(team) if team else ''
    pos_grp = _pos_group(position) if position else ''

    # Exact match
    if team and pos_grp:
        r = records.get((short, norm, norm_team, pos_grp))
        if r:
            return r
    else:
        for (gid, n, _t, _p), r in records.items():
            if gid == short and n == norm:
                return r

    # Fallback 1: first-initial + last-name, same team
    # (handles "J. Hagens" ↔ "James Hagens", spelling variants, etc.)
    init, last = _name_parts(norm)
    if not init or not last:
        return None
    candidates = []
    for (gid, n, t, p), r in records.items():
        if gid != short:
            continue
        if norm_team and t != norm_team:
            continue
        ri, rl = _name_parts(n)
        if ri and rl == last and ri[0] == init[0]:
            candidates.append(r)
    # Filter by position when known — prevents same-name players from stealing each other's records
    if pos_grp:
        pos_filtered = [c for c in candidates if c.position == pos_grp]
        if len(pos_filtered) == 1:
            return pos_filtered[0]
    elif len(candidates) == 1:
        return candidates[0]

    # Fallback 2: ignore team (handles mid-season trades where season_acc team ≠ playoff team)
    candidates = []
    for (gid, n, _t, _p), r in records.items():
        if gid != short:
            continue
        ri, rl = _name_parts(n)
        if ri and rl == last and ri[0] == init[0]:
            candidates.append(r)
    if pos_grp:
        pos_filtered = [c for c in candidates if c.position == pos_grp]
        if len(pos_filtered) == 1:
            return pos_filtered[0]
    elif len(candidates) == 1:
        return candidates[0]
    return None


def invalidate_cache() -> None:
    """Call this after adding new xlsx files so grades are recomputed."""
    global _CACHE, _RECORD_CACHE
    _CACHE = None
    _RECORD_CACHE = None
    if _CACHE_FILE.exists():
        _CACHE_FILE.unlink()
